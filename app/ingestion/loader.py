"""Document loaders — extract text from uploaded files and single web pages."""

import ipaddress
import logging
import socket
from pathlib import Path
from urllib.parse import urljoin, urlparse

import httpx
from bs4 import BeautifulSoup
from docx import Document
from pypdf import PdfReader

from app.config import settings

logger = logging.getLogger(__name__)

_BLOCKED_NETWORKS = (
    ipaddress.ip_network("0.0.0.0/8"),
    ipaddress.ip_network("10.0.0.0/8"),
    ipaddress.ip_network("127.0.0.0/8"),
    ipaddress.ip_network("169.254.0.0/16"),
    ipaddress.ip_network("172.16.0.0/12"),
    ipaddress.ip_network("192.168.0.0/16"),
    ipaddress.ip_network("224.0.0.0/4"),
    ipaddress.ip_network("240.0.0.0/4"),
    ipaddress.ip_network("::1/128"),
    ipaddress.ip_network("fc00::/7"),
    ipaddress.ip_network("fe80::/10"),
    ipaddress.ip_network("ff00::/8"),
)


def _is_blocked_ip(ip: ipaddress.IPv4Address | ipaddress.IPv6Address) -> bool:
    if ip.is_private or ip.is_loopback or ip.is_link_local or ip.is_multicast or ip.is_reserved:
        return True
    if ip.version == 6 and ip.ipv4_mapped is not None:
        return _is_blocked_ip(ip.ipv4_mapped)
    return any(ip in net for net in _BLOCKED_NETWORKS)


def _validate_public_url(url: str) -> str:
    parsed = urlparse(url)
    if parsed.scheme not in {"http", "https"} or not parsed.hostname:
        raise ValueError("URL not allowed")

    hostname = parsed.hostname
    try:
        infos = socket.getaddrinfo(hostname, parsed.port or (443 if parsed.scheme == "https" else 80))
    except socket.gaierror as exc:
        logger.warning("DNS resolution failed for host")
        raise ValueError("URL not allowed") from exc

    for info in infos:
        ip = ipaddress.ip_address(info[4][0])
        if _is_blocked_ip(ip):
            raise ValueError("URL not allowed")

    return url


async def _read_limited_body(response: httpx.Response) -> bytes:
    max_bytes = settings.web_fetch_max_bytes
    chunks: list[bytes] = []
    total = 0
    async for chunk in response.aiter_bytes():
        total += len(chunk)
        if total > max_bytes:
            raise ValueError("Could not fetch the page")
        chunks.append(chunk)
    return b"".join(chunks)


async def load_web_page(url: str) -> str:
    """Fetch a single public HTML page and return extracted visible text."""
    current = _validate_public_url(url.strip())
    timeout = httpx.Timeout(settings.web_fetch_timeout_seconds)
    headers = {"User-Agent": "AezaCodex/1.0 (+https://localhost; knowledge ingest)"}

    body = b""
    try:
        async with httpx.AsyncClient(follow_redirects=False, timeout=timeout, headers=headers) as client:
            for _ in range(5):
                async with client.stream("GET", current) as response:
                    if response.status_code in {301, 302, 303, 307, 308}:
                        location = response.headers.get("location")
                        if not location:
                            raise ValueError("Could not fetch the page")
                        current = _validate_public_url(urljoin(current, location))
                        continue

                    if response.status_code >= 400:
                        raise ValueError("Could not fetch the page")

                    body = await _read_limited_body(response)
                    break
            else:
                raise ValueError("Could not fetch the page")
    except ValueError:
        raise
    except httpx.TimeoutException as exc:
        raise ValueError("Timed out fetching the page") from exc
    except httpx.HTTPError as exc:
        raise ValueError("Could not fetch the page") from exc

    soup = BeautifulSoup(body, "html.parser")
    for tag in soup(["script", "style", "noscript", "svg"]):
        tag.decompose()

    text = soup.get_text(separator="\n", strip=True)
    if not text.strip():
        raise ValueError("No readable text found on that page")
    return text


def load_pdf(content: bytes) -> str:
    from io import BytesIO

    reader = PdfReader(BytesIO(content))
    parts = []
    for page in reader.pages:
        text = page.extract_text()
        if text:
            parts.append(text)
    return "\n".join(parts)


def load_docx(content: bytes) -> str:
    from io import BytesIO

    doc = Document(BytesIO(content))
    return "\n".join(p.text for p in doc.paragraphs if p.text.strip())


def load_txt(content: bytes) -> str:
    return content.decode("utf-8", errors="replace")


def _cell_text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _column_label(index: int) -> str:
    label = ""
    n = index + 1
    while n:
        n, rem = divmod(n - 1, 26)
        label = chr(65 + rem) + label
    return f"Column {label}"


def _is_empty_row(values: list[str]) -> bool:
    return not any(values)


def _sheet_records(sheet_name: str, rows: list[list]) -> list[str]:
    cleaned: list[list[str]] = []
    for row in rows:
        values = [_cell_text(v) for v in row]
        while values and not values[-1]:
            values.pop()
        if _is_empty_row(values):
            continue
        cleaned.append(values)

    if not cleaned:
        return []

    headers = cleaned[0]
    data_rows = cleaned[1:]
    if not data_rows:
        # Single row: treat as values with generic column names
        data_rows = [headers]
        headers = [_column_label(i) for i in range(len(headers))]
    else:
        headers = [h or _column_label(i) for i, h in enumerate(headers)]

    records: list[str] = []
    for row in data_rows:
        parts: list[str] = []
        width = max(len(headers), len(row))
        for i in range(width):
            header = headers[i] if i < len(headers) else _column_label(i)
            value = row[i] if i < len(row) else ""
            if not value:
                continue
            parts.append(f"{header}: {value}")
        if not parts:
            continue
        prefix = f"Sheet: {sheet_name} | " if sheet_name else ""
        records.append(prefix + " | ".join(parts))
    return records


def iter_excel_sheets(content: bytes, filename: str):
    """Yield (sheet_name, raw_rows) for each sheet in an Excel workbook."""
    suffix = Path(filename).suffix.lower()
    if suffix == ".xlsx":
        yield from _iter_xlsx_sheets(content)
    elif suffix == ".xls":
        yield from _iter_xls_sheets(content)
    else:
        raise ValueError(f"Unsupported file type: {suffix}")


def sheet_to_labeled_rows(sheet_name: str, rows: list[list]) -> list[str]:
    """Convert raw sheet rows into header: value record strings."""
    return _sheet_records(sheet_name, rows)


def extract_excel_records(content: bytes, filename: str) -> list[str]:
    """Turn Excel sheets into labeled row records (header: value)."""
    records: list[str] = []
    for name, rows in iter_excel_sheets(content, filename):
        records.extend(_sheet_records(name, rows))
    if not records:
        raise ValueError("No readable rows found in the spreadsheet")
    return records


def _iter_xlsx_sheets(content: bytes):
    from io import BytesIO

    from openpyxl import load_workbook

    workbook = load_workbook(BytesIO(content), data_only=True, read_only=True)
    try:
        for sheet in workbook.worksheets:
            rows = [list(row) for row in sheet.iter_rows(values_only=True)]
            yield sheet.title or "Sheet", rows
    finally:
        workbook.close()


def _iter_xls_sheets(content: bytes):
    import xlrd

    book = xlrd.open_workbook(file_contents=content)
    for sheet in book.sheets():
        rows = [
            [sheet.cell_value(r, c) for c in range(sheet.ncols)]
            for r in range(sheet.nrows)
        ]
        yield sheet.name or "Sheet", rows


def load_xlsx(content: bytes) -> str:
    return "\n".join(extract_excel_records(content, "file.xlsx"))


def load_xls(content: bytes) -> str:
    return "\n".join(extract_excel_records(content, "file.xls"))


LOADERS = {
    ".pdf": load_pdf,
    ".docx": load_docx,
    ".txt": load_txt,
    ".xlsx": load_xlsx,
    ".xls": load_xls,
}


def load_text(content: bytes, filename: str) -> str:
    """Dispatch to the correct loader based on file extension."""
    suffix = Path(filename).suffix.lower()
    loader = LOADERS.get(suffix)
    if loader is None:
        raise ValueError(f"Unsupported file type: {suffix}")
    return loader(content)
